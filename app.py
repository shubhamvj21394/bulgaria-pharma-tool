import streamlit as st
import pandas as pd
import numpy as np
import io, re, os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px
import plotly.graph_objects as go

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Bulgaria Pharma Pricing Tool",
    page_icon="💊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* Global */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

/* Background */
.stApp { background: #05080F; }
section[data-testid="stSidebar"] { background: #0D1525; }

/* Main header */
.main-header {
    background: linear-gradient(135deg, #0D1525 0%, #111D30 100%);
    border: 1px solid #1E3050;
    border-radius: 14px;
    padding: 28px 32px;
    margin-bottom: 24px;
    display: flex;
    align-items: center;
    gap: 18px;
}
.main-header h1 {
    font-size: 24px;
    font-weight: 800;
    color: #E8EEFF;
    margin: 0;
    letter-spacing: -0.5px;
}
.main-header p { color: #8BA3C7; font-size: 13px; margin: 4px 0 0; }

/* Metric cards */
.metric-row { display: flex; gap: 14px; margin-bottom: 20px; }
.metric-card {
    flex: 1;
    background: #111D30;
    border: 1px solid #1E3050;
    border-radius: 12px;
    padding: 18px 20px;
    text-align: center;
}
.metric-val {
    font-size: 32px;
    font-weight: 800;
    line-height: 1;
    font-family: 'Inter', sans-serif;
}
.metric-lbl {
    font-size: 11px;
    color: #8BA3C7;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    margin-top: 6px;
}
.mv-blue  { color: #3B82F6; }
.mv-green { color: #10B981; }
.mv-purple{ color: #8B5CF6; }
.mv-gold  { color: #F59E0B; }
.mv-red   { color: #EF4444; }
.mv-muted { color: #94A3B8; }

/* Upload boxes */
.upload-box {
    background: #0D1525;
    border: 1.5px dashed #1E3050;
    border-radius: 12px;
    padding: 20px;
    text-align: center;
    margin-bottom: 14px;
    transition: border-color .2s;
}
.upload-title { font-weight: 700; color: #E8EEFF; font-size: 14px; margin-bottom: 4px; }
.upload-sub   { font-size: 12px; color: #526A8A; }

/* Section headers */
.section-hdr {
    font-size: 11px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #8BA3C7;
    border-left: 3px solid #2563EB;
    padding-left: 10px;
    margin: 24px 0 14px;
}

/* Rule tags */
.rule-tag {
    display: inline-block;
    background: rgba(37,99,235,.1);
    border: 1px solid rgba(37,99,235,.2);
    border-radius: 6px;
    padding: 4px 10px;
    font-size: 11px;
    color: #93C5FD;
    margin: 3px 2px;
}

/* Status badge */
.badge-ok   { background:#064E3B; color:#6EE7B7; padding:3px 10px; border-radius:20px; font-size:11px; font-weight:600; }
.badge-warn { background:#451A03; color:#FCD34D; padding:3px 10px; border-radius:20px; font-size:11px; font-weight:600; }
.badge-err  { background:#450A0A; color:#FCA5A5; padding:3px 10px; border-radius:20px; font-size:11px; font-weight:600; }

/* Validation items */
.val-item {
    background: #0D1525;
    border: 1px solid #1E3050;
    border-radius: 8px;
    padding: 10px 14px;
    margin-bottom: 8px;
    font-size: 12.5px;
    color: #CBD5E1;
}

/* Step indicator */
.steps-row {
    display: flex;
    align-items: center;
    gap: 8px;
    margin-bottom: 20px;
    padding: 14px 16px;
    background: #0D1525;
    border: 1px solid #1E3050;
    border-radius: 10px;
}
.step-circle {
    width: 26px; height: 26px;
    border-radius: 50%;
    display: inline-flex; align-items: center; justify-content: center;
    font-size: 11px; font-weight: 700;
    flex-shrink: 0;
}
.step-active { background: #2563EB; color: #fff; }
.step-done   { background: #10B981; color: #fff; }
.step-idle   { background: #1E3050; color: #526A8A; }
.step-lbl    { font-size: 12px; color: #8BA3C7; }
.step-arr    { color: #1E3050; font-size: 12px; }

/* Plotly charts dark */
.js-plotly-plot .plotly { background: transparent !important; }

/* Download button styling */
.stDownloadButton > button {
    width: 100%;
    background: linear-gradient(135deg, #065F46, #047857) !important;
    color: white !important;
    border: 1px solid #10B981 !important;
    border-radius: 10px !important;
    padding: 12px !important;
    font-weight: 600 !important;
    font-size: 14px !important;
}
.stDownloadButton > button:hover {
    background: linear-gradient(135deg, #047857, #059669) !important;
    box-shadow: 0 4px 20px rgba(16,185,129,.3) !important;
}

/* Process button */
.stButton > button {
    width: 100%;
    background: linear-gradient(135deg, #1D4ED8, #2563EB) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 12px !important;
    font-weight: 600 !important;
    font-size: 14px !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #1E40AF, #1D4ED8) !important;
    box-shadow: 0 4px 20px rgba(37,99,235,.35) !important;
}
.stButton > button:disabled {
    background: #1E3050 !important;
    color: #526A8A !important;
}

/* Dataframe */
.dataframe { font-size: 12px !important; }
[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; }

/* File uploader */
[data-testid="stFileUploader"] {
    background: #0D1525 !important;
    border: 1.5px dashed #1E3050 !important;
    border-radius: 12px !important;
}
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RULES_FILE    = os.path.join(BASE_DIR, "Extraction_Rule.xlsx")
TEMPLATE_FILE = os.path.join(BASE_DIR, "Template.xlsx")

TEMPLATE_COLS = [
    "Action","Primary key_Pricing","PRICE_ID","Multiplication Factor","Country",
    "Active Ingredient","Brand Name","Company","Standard Form","Formulation",
    "Strength","Strength unit","Pack","Pack Unit","Fill","Fill Unit",
    "Effective Price Date","Currency (Local)","Manufacturer Price","Wholesale Price",
    "VAT","Retail Price without VAT","Retail Price","Price Launch Date","Launch Price",
    "Discontinued Date","Reimbursement","Reimbursement Comments","Hospital Product",
    "WHO ATC code","Combination product","Combination Strength","Combination Strength Unit",
    "Pack notes","Company Type","Pricing Strategy wrt lowest dose",
    "Pricing strategy across the dose","Local Brand Name","Local Company",
    "Local Pack Description","Source Name","File Date","Source Type"
]

DTYPE_MAP = {
    'Генеричен ЛП':                                                     'Generic',
    'Оригинален ЛП':                                                    'Originator',
    'Биоподобен ЛП':                                                    'Biosimilar',
    'Хибриден ЛП':                                                      'Hybrid',
    'Комбиниран ЛП с пълно досие':                                      'Combination - Full Dossier',
    'ЛП с добре установена употреба':                                   'Well-established Use',
    'Комбиниран ЛП, съдържащ вещества с добре установена употреба':     'Combination - Well-established Use',
}

PLOTLY_DARK = dict(
    plot_bgcolor='rgba(0,0,0,0)',
    paper_bgcolor='rgba(0,0,0,0)',
    font=dict(color='#8BA3C7', family='Inter', size=11),
    margin=dict(l=10, r=10, t=30, b=10),
)
GRID = dict(gridcolor='#1E3050', zerolinecolor='#1E3050')

# ── Helper functions ──────────────────────────────────────────────────────────
def parse_date(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})', s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    m = re.search(r'(\d{1,2})[./\-](\d{1,2})[./\-](\d{4})', s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    return s if s else None

def safe_num(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None
    try:
        return round(float(v), 4)
    except:
        return None

def safe_str(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None
    s = str(v).strip()
    return s if s else None

def get_file_date(raw):
    for i in range(min(5, len(raw))):
        for v in raw.iloc[i]:
            if 'Актуализация' in str(v):
                raw_date = str(v).split('Актуализация към')[-1].strip()
                return parse_date(raw_date)
    return datetime.now().strftime("%Y-%m-%d")

def find_header_rows(raw):
    for i in range(min(20, len(raw))):
        vals = [str(v) for v in raw.iloc[i]]
        if any('Международно непатентно' in v for v in vals):
            return i, i + 1
    raise ValueError("❌ Cannot find headers in file. Please check it is the correct Bulgarian pharmaceutical file.")

def process_source(raw, *, nat_col, upd_col, par_col, dtype_col,
                   source_label, file_date, h1, h2):
    data_start = h2 + 1
    rows = []
    for i in range(data_start, len(raw)):
        row = raw.iloc[i]
        inn = safe_str(row[0])
        if not inn:
            continue

        nat     = row[nat_col]
        name    = safe_str(row[2])
        company = safe_str(row[3])
        mfr_p   = safe_num(row[5])
        whl_p   = safe_num(row[10])
        rxv_p   = safe_num(row[15])
        ret_p   = safe_num(row[17])
        upd     = row[upd_col]
        par     = row[par_col]
        dtype   = safe_str(row[dtype_col]) if dtype_col is not None else None

        brand = name.split(',')[0].strip() if name else None
        par_str = safe_str(par)
        pack_notes = "Not Parallel Import" if par_str else "Parallel Import"

        if dtype:
            src_type = DTYPE_MAP.get(dtype, dtype)
        else:
            src_type = source_label

        try:
            price_id = str(int(float(nat))) if nat is not None and not (isinstance(nat, float) and np.isnan(nat)) else None
        except:
            price_id = safe_str(nat)

        rows.append({
            "Action": None, "Primary key_Pricing": None,
            "PRICE_ID": price_id,
            "Multiplication Factor": 1, "Country": "BULGARIA",
            "Active Ingredient": inn, "Brand Name": brand,
            "Company": company, "Standard Form": name,
            "Formulation": None, "Strength": None, "Strength unit": None,
            "Pack": None, "Pack Unit": None, "Fill": None, "Fill Unit": None,
            "Effective Price Date": parse_date(upd),
            "Currency (Local)": "BGN",
            "Manufacturer Price": mfr_p, "Wholesale Price": whl_p,
            "VAT": 20, "Retail Price without VAT": rxv_p, "Retail Price": ret_p,
            "Price Launch Date": None, "Launch Price": None, "Discontinued Date": None,
            "Reimbursement": None, "Reimbursement Comments": None,
            "Hospital Product": "No", "WHO ATC code": None,
            "Combination product": None, "Combination Strength": None,
            "Combination Strength Unit": None,
            "Pack notes": pack_notes, "Company Type": None,
            "Pricing Strategy wrt lowest dose": None,
            "Pricing strategy across the dose": None,
            "Local Brand Name": brand, "Local Company": company,
            "Local Pack Description": name,
            "Source Name": "NCPR", "File Date": file_date, "Source Type": src_type,
        })
    return pd.DataFrame(rows, columns=TEMPLATE_COLS)

def build_excel(df):
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    even_fill = PatternFill("solid", fgColor="EBF3FB")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")
    cell_font = Font(name="Calibri", size=9)
    cell_align = Alignment(vertical="center")
    thin  = Side(style="thin",   color="C5D5E8")
    thick = Side(style="medium", color="1E3A5F")
    cborder = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c_idx, col in enumerate(TEMPLATE_COLS, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = Border(left=thick, right=thick, top=thick, bottom=thick)
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(12, len(col)+2), 28)
    ws.row_dimensions[1].height = 32

    for r_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        fill = even_fill if r_idx % 2 == 0 else odd_fill
        for c_idx, col in enumerate(TEMPLATE_COLS, 1):
            val = row_data.get(col)
            if isinstance(val, float) and np.isnan(val):
                val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = cell_font; cell.fill = fill
            cell.alignment = cell_align; cell.border = cborder

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(TEMPLATE_COLS))}{len(df)+1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <div style="font-size:40px">💊</div>
  <div>
    <h1>Bulgaria Pharma Pricing Tool</h1>
    <p>Automated extraction, transformation &amp; standardisation of pharmaceutical pricing data &nbsp;·&nbsp;
       <strong style="color:#3B82F6">Bulgaria</strong> &nbsp;·&nbsp;
       <strong style="color:#3B82F6">BGN</strong> &nbsp;·&nbsp;
       <strong style="color:#3B82F6">NCPR</strong>
    </p>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Layout: Left sidebar + Main ───────────────────────────────────────────────
left_col, right_col = st.columns([1, 2.4], gap="large")

# ═══════════════════════════════════════════════════════════════════════════════
# LEFT COLUMN
# ═══════════════════════════════════════════════════════════════════════════════
with left_col:

    # Step indicator
    step = st.session_state.get("step", 1)
    def sc(n):
        if n < step: return "step-done"
        if n == step: return "step-active"
        return "step-idle"
    def st_txt(n):
        return "✓" if n < step else str(n)

    st.markdown(f"""
    <div class="steps-row">
      <span class="step-circle {sc(1)}">{st_txt(1)}</span>
      <span class="step-lbl">Upload</span>
      <span class="step-arr">›</span>
      <span class="step-circle {sc(2)}">{st_txt(2)}</span>
      <span class="step-lbl">Process</span>
      <span class="step-arr">›</span>
      <span class="step-circle {sc(3)}">{st_txt(3)}</span>
      <span class="step-lbl">Review</span>
      <span class="step-arr">›</span>
      <span class="step-circle {sc(4)}">{st_txt(4)}</span>
      <span class="step-lbl">Download</span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-hdr">Upload Source Files</div>', unsafe_allow_html=True)

    st.markdown('<div class="upload-title">📋 Appendix No. 4</div>', unsafe_allow_html=True)
    st.caption("Приложение № 4 на ПЛС · .xlsx / .xls")
    file_a4 = st.file_uploader("Appendix No. 4", type=["xlsx","xls"],
                                label_visibility="collapsed", key="upA4")

    st.markdown('<div class="upload-title" style="margin-top:10px">📊 Register of Marginal Prices</div>', unsafe_allow_html=True)
    st.caption("Регистър на пределните цени · .xlsx / .xls")
    file_pr = st.file_uploader("Register of Marginal Prices", type=["xlsx","xls"],
                                label_visibility="collapsed", key="upPr")

    both_uploaded = file_a4 is not None and file_pr is not None

    if both_uploaded and step == 1:
        st.session_state["step"] = 2

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    process_clicked = st.button(
        "⚡  Process Data",
        disabled=not both_uploaded,
        use_container_width=True,
    )

    # Download button (shown after processing)
    if "excel_bytes" in st.session_state:
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        st.download_button(
            label="⬇  Download Output Excel",
            data=st.session_state["excel_bytes"],
            file_name="Processed_Bulgaria_Drug_Data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # Extraction rules
    st.markdown('<div class="section-hdr" style="margin-top:20px">Extraction Rules (Backend)</div>',
                unsafe_allow_html=True)

    rules = [
        ("PRICE_ID", "Национален номер"),
        ("Active Ingredient", "INN (col A)"),
        ("Brand Name", "Drug name — before 1st comma"),
        ("Standard Form", "Full drug name string"),
        ("Company", "Притежател на разрешението"),
        ("Manufacturer Price", "цена under Цена на производител"),
        ("Wholesale Price", "цена under Утвърдена цена (Col K)"),
        ("Retail excl. VAT", "цена under Утвърдена цена (Col P)"),
        ("Retail Price", "общо с ДДС (Col R)"),
        ("Effective Date", "Дата актуализация → yyyy-mm-dd"),
        ("Pack Notes", "Parallel Import detection"),
        ("Source Type", "file label"),
        ("File Date", "Актуализация към → yyyy-mm-dd"),
        ("Country / Currency / VAT", "BULGARIA / BGN / 20"),
        ("Hospital Product", "No  ·  Source Name: NCPR  ·  Factor: 1"),
        ("Merge strategy", "CONCAT (no shared identifiers)"),
    ]
    for target, source in rules:
        st.markdown(
            f'<div style="font-size:12px;padding:3px 0;color:#CBD5E1">'
            f'<span style="color:#10B981">✓</span> '
            f'<strong style="color:#E8EEFF">{target}</strong>'
            f'<span style="color:#526A8A"> ← </span>'
            f'<span style="color:#8BA3C7">{source}</span></div>',
            unsafe_allow_html=True
        )

# ═══════════════════════════════════════════════════════════════════════════════
# RIGHT COLUMN — Processing + Dashboard
# ═══════════════════════════════════════════════════════════════════════════════
with right_col:

    if not both_uploaded:
        st.markdown("""
        <div style="display:flex;flex-direction:column;align-items:center;justify-content:center;
                    min-height:480px;background:#0D1525;border:1px solid #1E3050;
                    border-radius:14px;text-align:center;padding:40px">
          <div style="font-size:72px;opacity:.2;margin-bottom:18px">📈</div>
          <div style="font-family:'Inter',sans-serif;font-size:22px;font-weight:800;
                      color:#E8EEFF;opacity:.35;margin-bottom:10px">Processing Dashboard</div>
          <div style="font-size:14px;color:#526A8A;max-width:320px;line-height:1.7">
            Upload <strong style="color:#8BA3C7">Appendix No. 4</strong> and
            <strong style="color:#8BA3C7">Register of Marginal Prices</strong>,
            then click <strong style="color:#8BA3C7">Process Data</strong>.
          </div>
        </div>
        """, unsafe_allow_html=True)

    elif process_clicked or "results" in st.session_state:

        # ── Run processing ────────────────────────────────────────────────────
        if process_clicked:
            with st.spinner("🔄 Processing files — please wait…"):
                try:
                    prog = st.progress(0, text="Reading Appendix No. 4…")

                    raw4 = pd.read_excel(io.BytesIO(file_a4.read()), header=None)
                    prog.progress(15, "Reading Register of Marginal Prices…")
                    rawP = pd.read_excel(io.BytesIO(file_pr.read()), header=None)

                    prog.progress(30, "Detecting headers…")
                    h4a, h4b = find_header_rows(raw4)
                    hPa, hPb = find_header_rows(rawP)

                    file_date = get_file_date(raw4) or get_file_date(rawP)

                    prog.progress(45, "Processing Appendix No. 4…")
                    df_a4 = process_source(
                        raw4, nat_col=25, upd_col=21, par_col=22, dtype_col=26,
                        source_label="Appendix No.4",
                        file_date=file_date, h1=h4a, h2=h4b
                    )

                    prog.progress(60, "Processing Register of Marginal Prices…")
                    df_pr = process_source(
                        rawP, nat_col=26, upd_col=20, par_col=21, dtype_col=None,
                        source_label="Register of Marginal Prices",
                        file_date=file_date, h1=hPa, h2=hPb
                    )

                    prog.progress(75, "Merging datasets…")
                    final = pd.concat([df_a4, df_pr], ignore_index=True)

                    prog.progress(88, "Generating Excel output…")
                    excel_bytes = build_excel(final)

                    prog.progress(100, "✅ Complete!")

                    st.session_state["results"] = {
                        "final": final, "df_a4": df_a4, "df_pr": df_pr,
                        "file_date": file_date,
                    }
                    st.session_state["excel_bytes"] = excel_bytes
                    st.session_state["step"] = 4
                    st.rerun()

                except Exception as e:
                    st.error(f"**Processing Error:** {e}")
                    st.stop()

        # ── Display Dashboard ─────────────────────────────────────────────────
        if "results" in st.session_state:
            res   = st.session_state["results"]
            final = res["final"]
            df_a4 = res["df_a4"]
            df_pr = res["df_pr"]
            fd    = res["file_date"]

            st.success(f"✅  **{len(final):,} records** processed successfully from both files  ·  File Date: `{fd}`")

            # ── KPI Metrics ───────────────────────────────────────────────────
            st.markdown('<div class="section-hdr">Summary Metrics</div>', unsafe_allow_html=True)

            c1,c2,c3,c4,c5,c6 = st.columns(6)
            kpis = [
                (c1, len(final), "Total Records", "mv-blue"),
                (c2, len(df_a4), "Appendix No.4", "mv-green"),
                (c3, len(df_pr), "Register Rows", "mv-purple"),
                (c4, len(final), "Merged Records", "mv-gold"),
                (c5, int(final["Manufacturer Price"].isna().sum()), "Missing Prices", "mv-red"),
                (c6, fd, "File Date", "mv-muted"),
            ]
            for col, val, lbl, cls in kpis:
                with col:
                    display = f"{val:,}" if isinstance(val, int) else str(val)
                    st.markdown(f"""
                    <div class="metric-card">
                      <div class="metric-val {cls}">{display}</div>
                      <div class="metric-lbl">{lbl}</div>
                    </div>""", unsafe_allow_html=True)

            # ── Charts ────────────────────────────────────────────────────────
            st.markdown('<div class="section-hdr">Analytics Charts</div>', unsafe_allow_html=True)

            chart_c1, chart_c2 = st.columns(2)

            # Chart 1: Top manufacturers
            with chart_c1:
                mfr_counts = (
                    final["Company"].dropna()
                    .apply(lambda x: x.split(',')[0].strip()[:40])
                    .value_counts().head(12).reset_index()
                )
                mfr_counts.columns = ["Manufacturer", "Count"]
                fig1 = px.bar(
                    mfr_counts, x="Count", y="Manufacturer",
                    orientation="h", title="Top Manufacturers by Record Count",
                    color_discrete_sequence=["#3B82F6"],
                )
                fig1.update_layout(**PLOTLY_DARK)
                fig1.update_xaxes(**GRID)
                fig1.update_yaxes(**GRID)
                fig1.update_traces(marker_line_width=0)
                fig1.update_yaxes(categoryorder="total ascending")
                st.plotly_chart(fig1, use_container_width=True)

            # Chart 2: Price distribution donut
            with chart_c2:
                prices = final["Retail Price"].dropna()
                labels = ["0–10","10–50","50–200","200–500","500+"]
                counts = [
                    (prices < 10).sum(),
                    ((prices >= 10) & (prices < 50)).sum(),
                    ((prices >= 50) & (prices < 200)).sum(),
                    ((prices >= 200) & (prices < 500)).sum(),
                    (prices >= 500).sum(),
                ]
                fig2 = go.Figure(go.Pie(
                    labels=[f"BGN {l}" for l in labels],
                    values=counts,
                    hole=0.45,
                    marker_colors=["#3B82F6","#6366F1","#F59E0B","#10B981","#EF4444"],
                ))
                fig2.update_layout(title="Retail Price Distribution (BGN)", **PLOTLY_DARK)
                fig2.update_traces(textfont_color="#E8EEFF")
                st.plotly_chart(fig2, use_container_width=True)

            chart_c3, chart_c4 = st.columns(2)

            # Chart 3: Top priced drugs
            with chart_c3:
                top10 = (
                    final[["Brand Name","Retail Price"]].dropna(subset=["Retail Price"])
                    .sort_values("Retail Price", ascending=False).head(10)
                )
                top10["Brand Name"] = top10["Brand Name"].apply(lambda x: str(x)[:30] if x else "")
                fig3 = px.bar(
                    top10, x="Retail Price", y="Brand Name",
                    orientation="h", title="Top 10 Highest-Priced Drugs (BGN)",
                    color_discrete_sequence=["#F59E0B"],
                )
                fig3.update_layout(**PLOTLY_DARK)
                fig3.update_xaxes(**GRID)
                fig3.update_yaxes(**GRID)
                fig3.update_traces(marker_line_width=0)
                fig3.update_yaxes(categoryorder="total ascending")
                fig3.update_xaxes(tickprefix="BGN ")
                st.plotly_chart(fig3, use_container_width=True)

            # Chart 4: Source type breakdown
            with chart_c4:
                src_counts = final["Source Type"].fillna("Unknown").value_counts().reset_index()
                src_counts.columns = ["Source Type", "Count"]
                fig4 = px.pie(
                    src_counts, names="Source Type", values="Count",
                    title="Source Type Breakdown",
                    hole=0.4,
                    color_discrete_sequence=px.colors.qualitative.Bold,
                )
                fig4.update_layout(**PLOTLY_DARK)
                fig4.update_traces(textfont_color="#E8EEFF")
                st.plotly_chart(fig4, use_container_width=True)

            # Price comparison line chart
            sample = (
                final[["Brand Name","Manufacturer Price","Retail Price"]]
                .dropna().head(30)
            )
            fig5 = go.Figure()
            fig5.add_trace(go.Scatter(
                x=sample["Brand Name"].apply(lambda x: str(x)[:22]),
                y=sample["Manufacturer Price"],
                name="Manufacturer Price", line=dict(color="#3B82F6", width=2),
                mode="lines+markers", marker=dict(size=4),
            ))
            fig5.add_trace(go.Scatter(
                x=sample["Brand Name"].apply(lambda x: str(x)[:22]),
                y=sample["Retail Price"],
                name="Retail Price", line=dict(color="#10B981", width=2),
                mode="lines+markers", marker=dict(size=4),
            ))
            fig5.update_layout(
                title="Manufacturer vs Retail Price (First 30 Records)",
                **PLOTLY_DARK,
                legend=dict(font=dict(color="#8BA3C7"), bgcolor="rgba(0,0,0,0)"),
                height=300,
            )
            fig5.update_xaxes(tickangle=45, **GRID)
            fig5.update_yaxes(**GRID)
            st.plotly_chart(fig5, use_container_width=True)

            # ── Validation Report ─────────────────────────────────────────────
            st.markdown('<div class="section-hdr">Validation &amp; Quality Report</div>',
                        unsafe_allow_html=True)

            v_col1, v_col2, v_col3 = st.columns(3)

            dup_ids = final[final.duplicated("PRICE_ID", keep=False)]
            miss_price = final[final["Manufacturer Price"].isna()]
            miss_id = final[final["PRICE_ID"].isna()]

            with v_col1:
                if len(miss_id) == 0:
                    st.markdown('<span class="badge-ok">✓ No missing PRICE_IDs</span>',
                                unsafe_allow_html=True)
                else:
                    st.markdown(f'<span class="badge-err">⚠ {len(miss_id)} missing PRICE_IDs</span>',
                                unsafe_allow_html=True)

            with v_col2:
                if len(miss_price) == 0:
                    st.markdown('<span class="badge-ok">✓ All prices present</span>',
                                unsafe_allow_html=True)
                else:
                    st.markdown(f'<span class="badge-warn">⚠ {len(miss_price)} missing prices</span>',
                                unsafe_allow_html=True)

            with v_col3:
                if len(dup_ids) == 0:
                    st.markdown('<span class="badge-ok">✓ No duplicate PRICE_IDs</span>',
                                unsafe_allow_html=True)
                else:
                    st.markdown(f'<span class="badge-warn">⚠ {len(dup_ids)} duplicate PRICE_IDs</span>',
                                unsafe_allow_html=True)

            st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

            # ── Data Tables ───────────────────────────────────────────────────
            st.markdown('<div class="section-hdr">Data Preview &amp; Quality Tables</div>',
                        unsafe_allow_html=True)

            tab1, tab2, tab3, tab4 = st.tabs([
                "📄 Output Preview",
                "🔴 Duplicates",
                "⚠️ Missing Values",
                "💰 Top Priced",
            ])

            with tab1:
                preview_cols = ["PRICE_ID","Active Ingredient","Brand Name","Company",
                                "Manufacturer Price","Wholesale Price","Retail Price without VAT",
                                "Retail Price","Effective Price Date","Pack notes","Source Type"]
                st.dataframe(
                    final[preview_cols].head(100),
                    use_container_width=True,
                    height=380,
                )
                st.caption(f"Showing first 100 of {len(final):,} records · {len(TEMPLATE_COLS)} columns total")

            with tab2:
                if len(dup_ids) == 0:
                    st.success("✅ No duplicate PRICE_IDs detected across both files.")
                else:
                    st.warning(f"{len(dup_ids)} rows share duplicate PRICE_IDs.")
                    st.dataframe(
                        dup_ids[["PRICE_ID","Brand Name","Company","Retail Price","Source Type"]],
                        use_container_width=True, height=300,
                    )

            with tab3:
                miss_df = final[final[["PRICE_ID","Active Ingredient","Manufacturer Price"]].isna().any(axis=1)]
                if len(miss_df) == 0:
                    st.success("✅ All required fields are populated — no missing values.")
                else:
                    st.warning(f"{len(miss_df)} records have missing required values.")
                    st.dataframe(
                        miss_df[["PRICE_ID","Active Ingredient","Brand Name","Manufacturer Price","Source Type"]],
                        use_container_width=True, height=300,
                    )

            with tab4:
                top_priced = (
                    final[["Brand Name","Active Ingredient","Manufacturer Price",
                           "Wholesale Price","Retail Price","Source Type"]]
                    .dropna(subset=["Retail Price"])
                    .sort_values("Retail Price", ascending=False)
                    .head(25)
                )
                st.dataframe(top_priced, use_container_width=True, height=380)

            # Final download reminder
            st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
            st.info(
                f"📥  **Output ready** — {len(final):,} records · 43 columns · "
                f"Click **⬇ Download Output Excel** in the left panel to save the file."
            )
